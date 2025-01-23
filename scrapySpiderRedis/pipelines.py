# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

# useful for handling different item types with a single interface
# from itemadapter import ItemAdapter

# class ScrapyspiderredisPipeline:
#     def process_item(self, item, spider):
#         return item

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from scrapy.exceptions import DropItem
    
import pymongo
import pymysql
import time
from functools import wraps
## 数据写入到mongodb
class MongoPipeline(object):
    def __init__(self, mongo_uri, mongo_db):
        self.mongo_uri = mongo_uri
        self.mongo_db = mongo_db
    
    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            mongo_uri=crawler.settings.get('MONGO_URI'),
            mongo_db=crawler.settings.get('MONGO_DB')
        )
    
    def open_spider(self, spider):
        self.client = pymongo.MongoClient(self.mongo_uri)
        self.db = self.client[self.mongo_db]
    
    def process_item(self, item, spider):
        try:
            name = item.collection
            self.db[name].insert_one(dict(item))  # 插入单个文档
        except Exception as e:
            spider.logger.info(f"Error inserting item into MongoDB: {e}")
        return item
    
    def close_spider(self, spider):
        self.client.close()

## 数据写入到mysql,通过pipelines里=前面的字段进行自动填充
class MysqlPipeline():
    def __init__(self, host, database, user, password, port):
        self.host = host
        self.database = database
        self.user = user
        self.password = password
        self.port = port

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            host=crawler.settings.get('MYSQL_HOST'),
            database=crawler.settings.get('MYSQL_DATABASE'),
            user=crawler.settings.get('MYSQL_USER'),
            password=crawler.settings.get('MYSQL_PASSWORD'),
            port=crawler.settings.get('MYSQL_PORT'),
        )

    def open_spider(self, spider):
        self.connect_to_db()

    def close_spider(self, spider):
        self.db.close()

    def connect_to_db(self):
        self.db = pymysql.connect(host=self.host, user=self.user, password=self.password, database=self.database, charset='utf8',
                                  port=self.port)
        self.cursor = self.db.cursor()

    def retry_on_disconnect(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            self = args[0]
            max_retries = 5
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except pymysql.MySQLError as e:
                    if e.args[0] in (2006, 2013, 2014):  # MySQL server has gone away, Lost connection to MySQL server during query, Commands out of sync
                        time.sleep(2 ** attempt)  # Exponential backoff
                        self.connect_to_db()
                    else:
                        raise
        return wrapper

    def clean_data(self, item: dict) -> bool:
        return True

    @retry_on_disconnect
    def process_item(self, item, spider):
        try:
            if self.clean_data(item):
                table_name = item.table

                fields = item.fields.keys()
                field_names = ', '.join(fields)
                field_values = ', '.join(['%s'] * len(fields))

                sql = f"INSERT INTO {table_name} ({field_names}) VALUES ({field_values})"
                params = [item.get(field) for field in fields]

                self.cursor.execute(sql, params)
                self.db.commit()
        except Exception as e:
            spider.logger.info(f"Error occurred: {e}")
            self.db.rollback()
        return item
    
class ExcelPipeline():
    """_summary_
    爬取数据下载为excel
    """
    def __init__(self, file_name):
        self.file_name = file_name

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            file_name=crawler.settings.get('EXCEL_FILE_NAME', 'output.xlsx')
        )

    def open_spider(self, spider):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def close_spider(self, spider):
        self.workbook.save(self.file_name)

    def clean_data(self, item: dict) -> bool:
        # Implement your data cleaning logic here
        return True

    def process_item(self, item, spider):
        try:
            if self.clean_data(item):
                if self.worksheet.max_row == 1:  # Write headers only once
                    headers = list(item.keys())
                    self.worksheet.append(headers)
                
                row = [item.get(field, '') for field in item.keys()]
                self.worksheet.append(row)
        except Exception as e:
            spider.logger.info(f"Error occurred: {e}")
            raise DropItem(f"Error writing to Excel: {e}")
        return item
